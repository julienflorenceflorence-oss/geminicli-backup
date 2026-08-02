import os
import sys
import json
import sqlite3
import logging
from datetime import datetime, timedelta
import urllib.request
import urllib.parse
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import shutil

# Import optionnel de pandas et openpyxl pour l'export Excel stylisé.
# Si non installés, on affiche un message d'erreur clair.
try:
    import pandas as pd
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False

# Configuration globale du projet et chemins
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
CONFIG_FILE = os.path.join(SCRIPT_DIR, "config.json")

# Charger la configuration
def load_config():
    if not os.path.exists(CONFIG_FILE):
        print(f"Erreur : Le fichier de configuration '{CONFIG_FILE}' est introuvable.")
        sys.exit(1)
    with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
        return json.load(f)

config = load_config()

# Setup des chemins absolus à partir de la config
DATABASE_PATH = os.path.abspath(os.path.join(SCRIPT_DIR, config["paths"]["database_path"]))
EXCEL_OUTPUT_DIR = os.path.abspath(os.path.join(SCRIPT_DIR, config["paths"]["excel_output_dir"]))
LOG_OUTPUT_DIR = os.path.abspath(os.path.join(SCRIPT_DIR, config["paths"]["log_output_dir"]))

# S'assurer que les dossiers de destination existent
os.makedirs(os.path.dirname(DATABASE_PATH), exist_ok=True)
os.makedirs(EXCEL_OUTPUT_DIR, exist_ok=True)
os.makedirs(os.path.join(EXCEL_OUTPUT_DIR, "Archives"), exist_ok=True)
os.makedirs(LOG_OUTPUT_DIR, exist_ok=True)

# Configuration de la journalisation (Logs)
log_file = os.path.join(LOG_OUTPUT_DIR, "dpe_watcher.log")
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler(log_file, encoding='utf-8'),
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger("DPEWatcher")

# Initialisation de la base de données historique SQLite
def init_db():
    logger.info(f"Initialisation de la base de données historique : {DATABASE_PATH}")
    conn = sqlite3.connect(DATABASE_PATH)
    cursor = conn.cursor()
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS processed_dpes (
            numero_dpe TEXT PRIMARY KEY,
            date_recup TEXT NOT NULL,
            date_etablissement_dpe TEXT,
            code_postal_brut TEXT,
            nom_commune_brut TEXT
        )
    """)
    conn.commit()
    conn.close()

# Vérifier si un DPE a déjà été traité
def is_dpe_processed(numero_dpe):
    conn = sqlite3.connect(DATABASE_PATH)
    cursor = conn.cursor()
    cursor.execute("SELECT 1 FROM processed_dpes WHERE numero_dpe = ?", (numero_dpe,))
    result = cursor.fetchone()
    conn.close()
    return result is not None

# Ajouter des DPE à l'historique
def mark_dpes_as_processed(dpe_records):
    if not dpe_records:
        return
    conn = sqlite3.connect(DATABASE_PATH)
    cursor = conn.cursor()
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    # Préparation des tuples pour insertion en masse
    insert_data = [
        (
            r.get("numero_dpe"),
            now_str,
            r.get("date_etablissement_dpe"),
            str(r.get("code_postal_brut")),
            r.get("nom_commune_brut")
        )
        for r in dpe_records if r.get("numero_dpe")
    ]
    
    cursor.executemany("""
        INSERT OR IGNORE INTO processed_dpes 
        (numero_dpe, date_recup, date_etablissement_dpe, code_postal_brut, nom_commune_brut)
        VALUES (?, ?, ?, ?, ?)
    """, insert_data)
    conn.commit()
    conn.close()
    logger.info(f"{len(insert_data)} nouveaux DPE enregistrés dans l'historique local SQLite.")

# Interroger l'API ADEME
def fetch_new_dpes_from_api():
    if not PANDAS_AVAILABLE:
        logger.error("Les bibliothèques pandas ou openpyxl ne sont pas disponibles. Exécutez 'pip install -r requirements.txt'")
        sys.exit(1)
        
    codes_postaux = config["filtering"]["codes_postaux"]
    days_to_check = config["ademe_api"]["days_to_check"]
    datasets = config["ademe_api"]["datasets"]
    
    if not codes_postaux:
        logger.warning("Aucun code postal n'est configuré dans config.json. La recherche sera vide.")
        return []
    
    # Calcul des dates de filtrage
    end_date = datetime.now()
    start_date = end_date - timedelta(days=days_to_check)
    start_date_str = start_date.strftime("%Y-%m-%d")
    end_date_str = end_date.strftime("%Y-%m-%d")
    
    # Construction de la clause Lucene
    # Exemple: code_postal_brut:(11000 OR 11100) AND date_etablissement_dpe:[2026-06-25 TO 2026-07-02]
    cp_query = " OR ".join(codes_postaux)
    lucene_query = f"code_postal_brut:({cp_query}) AND date_etablissement_dpe:[{start_date_str} TO {end_date_str}]"
    logger.info(f"Filtre de recherche (Lucene): {lucene_query}")
    
    selected_fields = (
        "nom_commune_brut,adresse_brut,code_postal_brut,date_etablissement_dpe,"
        "numero_dpe,type_batiment,etiquette_dpe,periode_construction,surface_habitable_logement"
    )
    
    new_dpes = []
    
    for dataset in datasets:
        dataset_id = dataset["id"]
        dataset_name = dataset["name"]
        logger.info(f"Interrogation du jeu de données : {dataset_name} ({dataset_id})")
        
        # URL de départ pour l'API DataFair
        url = f"https://data.ademe.fr/data-fair/api/v1/datasets/{dataset_id}/lines"
        params = {
            "size": 100,  # Récupérer par lots de 100
            "qs": lucene_query,
            "select": selected_fields
        }
        
        query_string = urllib.parse.urlencode(params)
        next_url = f"{url}?{query_string}"
        
        dataset_count = 0
        dataset_new_count = 0
        
        while next_url:
            logger.info(f"Requête API : {next_url}")
            try:
                req = urllib.request.Request(
                    next_url, 
                    headers={"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) DPEWatcher/1.0"}
                )
                with urllib.request.urlopen(req, timeout=config["ademe_api"]["request_timeout_seconds"]) as response:
                    data = json.loads(response.read().decode('utf-8'))
                    
                    results = data.get("results", [])
                    dataset_count += len(results)
                    
                    for r in results:
                        num_dpe = r.get("numero_dpe")
                        if num_dpe and not is_dpe_processed(num_dpe):
                            new_dpes.append(r)
                            dataset_new_count += 1
                    
                    # Récupération de la page suivante
                    next_url = data.get("next")
            except Exception as e:
                logger.error(f"Erreur lors de la requête API sur {dataset_name} : {e}")
                break
                
        logger.info(f"Terminé pour {dataset_name} : {dataset_count} DPE analysés, {dataset_new_count} nouveaux DPE identifiés.")
        
    return new_dpes

# Génération et stylisation du fichier Excel "Édition Prestige"
def generate_excel(new_dpes):
    if not new_dpes:
        logger.info("Aucun nouveau DPE à exporter.")
        return None
        
    # Conversion en DataFrame
    df = pd.DataFrame(new_dpes)
    
    # Réarrangement et renommage des colonnes demandées
    columns_mapping = {
        "nom_commune_brut": "Commune",
        "adresse_brut": "Adresse",
        "code_postal_brut": "Code Postal",
        "date_etablissement_dpe": "Date Établissement",
        "numero_dpe": "N° DPE",
        "type_batiment": "Type Bâtiment",
        "etiquette_dpe": "Classe DPE",
        "periode_construction": "Période Construction",
        "surface_habitable_logement": "Surface Habitable (m²)"
    }
    
    # S'assurer que toutes les colonnes requises existent dans le DataFrame
    for col in columns_mapping.keys():
        if col not in df.columns:
            df[col] = None
            
    df = df[list(columns_mapping.keys())]
    df.rename(columns=columns_mapping, inplace=True)
    
    # Tri par date décroissante
    df["Date Établissement"] = pd.to_datetime(df["Date Établissement"], errors='coerce')
    df.sort_values(by="Date Établissement", ascending=False, inplace=True)
    
    # Formater la date en chaîne propre pour l'Excel
    df["Date Établissement"] = df["Date Établissement"].dt.strftime('%d/%m/%Y')
    df.fillna("", inplace=True)
    
    # Gestion du versioning binaire du fichier actif Excel
    excel_filename = "DPE_Nouveaux.xlsx"
    excel_path = os.path.join(EXCEL_OUTPUT_DIR, excel_filename)
    
    # Si le fichier actif existe déjà, on l'archive
    if os.path.exists(excel_path):
        archive_timestamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
        archive_filename = f"DPE_Nouveaux_{archive_timestamp}.xlsx"
        archive_path = os.path.join(EXCEL_OUTPUT_DIR, "Archives", archive_filename)
        shutil.copy2(excel_path, archive_path)
        logger.info(f"Fichier actif précédent archivé sous : {archive_path}")
        
    # Création du fichier Excel stylisé "Édition Prestige"
    logger.info(f"Génération du fichier Excel stylisé : {excel_path}")
    writer = pd.ExcelWriter(excel_path, engine='openpyxl')
    df.to_excel(writer, index=False, sheet_name="DPE Nouveaux")
    
    # Application du style Prestige avec openpyxl
    workbook = writer.book
    worksheet = writer.sheets["DPE Nouveaux"]
    
    # Définition des couleurs de la charte Prestige
    fill_header = PatternFill(start_color="0F1115", end_color="0F1115", fill_type="solid") # Noir profond
    font_header = Font(name="Segoe UI", size=11, bold=True, color="D4AF37") # Texte Doré
    font_body = Font(name="Segoe UI", size=10, color="333333")
    fill_zebra = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid") # Légère teinte grise
    
    border_thin = Side(border_style="thin", color="E5E7EB")
    border_header_bottom = Side(border_style="medium", color="D4AF37") # Ligne inférieure dorée
    
    border_cell = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)
    border_header = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_header_bottom)
    
    # Alignements
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    
    # Styliser l'en-tête
    for col_idx in range(1, len(df.columns) + 1):
        cell = worksheet.cell(row=1, column=col_idx)
        cell.fill = fill_header
        cell.font = font_header
        cell.alignment = align_center
        cell.border = border_header
        
    # Styliser les lignes de données
    for row_idx in range(2, len(df) + 2):
        is_even = row_idx % 2 == 0
        for col_idx in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            cell.font = font_body
            cell.border = border_cell
            
            # Effet zèbre discret
            if is_even:
                cell.fill = fill_zebra
                
            # Alignements spécifiques aux colonnes
            col_name = df.columns[col_idx - 1]
            if col_name in ["Code Postal", "Date Établissement", "Classe DPE"]:
                cell.alignment = align_center
            elif col_name in ["N° DPE"]:
                cell.alignment = align_center
            elif col_name in ["Surface Habitable (m²)"]:
                cell.alignment = align_right
                # Formater les nombres si possible
                if cell.value != "":
                    try:
                        cell.value = float(cell.value)
                        cell.number_format = '0.0'
                    except ValueError:
                        pass
            else:
                cell.alignment = align_left

    # Ajustement automatique de la largeur des colonnes
    for col in worksheet.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            # Remplacement des sauts de ligne pour calculer la largeur
            val_str = str(cell.value or '').split('\n')[0]
            if len(val_str) > max_len:
                max_len = len(val_str)
        # Largeur min de 12, max de 40 pour l'adresse
        worksheet.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 40)
        
    # Figer les en-têtes
    worksheet.freeze_panes = "A2"
    
    # Sauvegarde
    writer.close()
    logger.info("Fichier Excel généré et stylisé avec succès.")
    return excel_path

# Envoi de l'email avec le fichier Excel en pièce jointe
def send_email(attachment_path, num_records):
    email_config = config["email"]
    
    # S'il n'y a pas d'enregistrements et qu'on ne veut pas envoyer d'email vide
    if num_records == 0 and not email_config["send_email_if_empty"]:
        logger.info("Aucun nouveau DPE. Envoi d'email ignoré (conformément à la configuration).")
        return
        
    subject = f"{email_config['subject_prefix']} Rapport quotidien - {num_records} nouveau(x) DPE"
    
    # Création du message
    msg = MIMEMultipart()
    msg['From'] = email_config['from_email']
    msg['To'] = ", ".join(email_config['to_emails'])
    msg['Subject'] = subject
    
    # Corps de l'e-mail au format HTML avec un léger design élégant
    date_now = datetime.now().strftime("%d/%m/%Y à %H:%M")
    if num_records > 0:
        html_body = f"""
        <html>
        <body style="font-family: Arial, sans-serif; color: #333; line-height: 1.6;">
            <div style="max-width: 600px; margin: 0 auto; border: 1px solid #e5e7eb; border-radius: 8px; overflow: hidden; box-shadow: 0 4px 6px rgba(0,0,0,0.05);">
                <div style="background-color: #0f1115; padding: 20px; text-align: center; border-bottom: 3px solid #d4af37;">
                    <h2 style="color: #d4af37; margin: 0; font-weight: normal; letter-spacing: 1px;">DPE WATCHER</h2>
                    <p style="color: #a3a3a3; margin: 5px 0 0 0; font-size: 13px;">Rapport automatique de veille ADEME</p>
                </div>
                <div style="padding: 25px; background-color: #ffffff;">
                    <p>Bonjour,</p>
                    <p>L'analyse quotidienne de l'API ADEME a été effectuée avec succès le <strong>{date_now}</strong>.</p>
                    <div style="background-color: #f9fafb; border-left: 4px solid #d4af37; padding: 15px; margin: 20px 0; border-radius: 0 4px 4px 0;">
                        <span style="font-size: 24px; font-weight: bold; color: #0f1115;">{num_records}</span> 
                        <span style="font-size: 16px; color: #4b5563; margin-left: 5px;">nouveau(x) DPE identifié(s) sur vos codes postaux cibles.</span>
                    </div>
                    <p>Le fichier Excel récapitulatif contenant les nouvelles données est joint à ce message.</p>
                    <p style="font-size: 12px; color: #6b7280; margin-top: 30px; border-top: 1px solid #f3f4f6; padding-top: 15px;">
                        <em>Ce message a été généré automatiquement par votre robot de surveillance DPE.</em>
                    </p>
                </div>
            </div>
        </body>
        </html>
        """
    else:
        html_body = f"""
        <html>
        <body style="font-family: Arial, sans-serif; color: #333; line-height: 1.6;">
            <div style="max-width: 600px; margin: 0 auto; border: 1px solid #e5e7eb; border-radius: 8px; overflow: hidden;">
                <div style="background-color: #0f1115; padding: 20px; text-align: center; border-bottom: 3px solid #d4af37;">
                    <h2 style="color: #d4af37; margin: 0; font-weight: normal; letter-spacing: 1px;">DPE WATCHER</h2>
                </div>
                <div style="padding: 25px; background-color: #ffffff;">
                    <p>Bonjour,</p>
                    <p>L'analyse quotidienne de l'API ADEME a été effectuée avec succès le <strong>{date_now}</strong>.</p>
                    <p><strong>Aucun nouveau DPE</strong> n'a été détecté pour vos codes postaux cibles durant les derniers jours.</p>
                    <p style="font-size: 12px; color: #6b7280; margin-top: 30px; border-top: 1px solid #f3f4f6; padding-top: 15px;">
                        <em>Ce message a été généré automatiquement par votre robot de surveillance DPE.</em>
                    </p>
                </div>
            </div>
        </body>
        </html>
        """
        
    msg.attach(MIMEText(html_body, 'html'))
    
    # Pièce jointe
    if attachment_path and os.path.exists(attachment_path):
        filename = os.path.basename(attachment_path)
        with open(attachment_path, "rb") as attachment:
            part = MIMEBase("application", "octet-stream")
            part.set_payload(attachment.read())
            encoders.encode_base64(part)
            part.add_header(
                "Content-Disposition",
                f"attachment; filename= {filename}",
            )
            msg.attach(part)
            
    # Connexion et envoi SMTP
    try:
        logger.info(f"Connexion au serveur SMTP {email_config['smtp_server']}:{email_config['smtp_port']}...")
        if email_config["use_ssl"]:
            server = smtplib.SMTP_SSL(email_config['smtp_server'], email_config['smtp_port'])
        else:
            server = smtplib.SMTP(email_config['smtp_server'], email_config['smtp_port'])
            if email_config["use_tls"]:
                server.starttls()
                
        if email_config['smtp_username'] and email_config['smtp_password']:
            server.login(email_config['smtp_username'], email_config['smtp_password'])
            
        logger.info("Envoi de l'email...")
        server.sendmail(
            email_config['from_email'],
            email_config['to_emails'],
            msg.as_string()
        )
        server.quit()
        logger.info("Email envoyé avec succès !")
    except Exception as e:
        logger.error(f"Échec de l'envoi de l'e-mail : {e}")

# Point d'entrée principal
def main():
    logger.info("=== DÉBUT DE L'EXÉCUTION DU DPE WATCHER ===")
    
    try:
        init_db()
        new_dpes = fetch_new_dpes_from_api()
        
        if new_dpes:
            logger.info(f"{len(new_dpes)} nouveaux DPE uniques découverts au total.")
            excel_file = generate_excel(new_dpes)
            
            # Enregistrement dans l'historique pour ne plus les retraiter
            mark_dpes_as_processed(new_dpes)
            
            # Envoi de l'email
            send_email(excel_file, len(new_dpes))
        else:
            logger.info("Aucun nouveau DPE trouvé aujourd'hui.")
            # Si configuré pour envoyer même si vide
            if config["email"]["send_email_if_empty"]:
                send_email(None, 0)
                
    except Exception as e:
        logger.critical(f"Erreur critique lors de l'exécution : {e}", exc_info=True)
        
    logger.info("=== FIN DE L'EXÉCUTION DO DPE WATCHER ===")

if __name__ == "__main__":
    main()
